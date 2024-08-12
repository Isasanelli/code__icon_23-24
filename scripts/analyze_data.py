import pandas as pd
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def load_dataset(file_path: str) -> pd.DataFrame:
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Il file {file_path} non esiste. Assicurati che il percorso sia corretto.")
    return pd.read_csv(file_path)

def convert_to_categorical(dataset: pd.DataFrame, columns: list) -> pd.DataFrame:
    for col in columns:
        dataset[col] = dataset[col].astype('category')
    return dataset

def generate_statistics(dataset: pd.DataFrame) -> pd.DataFrame:
    # Escludi colonne con testo lungo per evitare distorsioni nei risultati
    columns_to_exclude = ['description']
    dataset = dataset.drop(columns=columns_to_exclude)
    
    # Elimina colonne con troppi valori mancanti
    dataset = dataset.dropna(axis=1, thresh=int(0.5 * len(dataset)))

    # Genera statistiche per colonne categoriche e numeriche separatamente
    numerical_stats = dataset.describe(include=[float, int])
    categorical_stats = dataset.describe(include=['category'])
    
    # Combina le statistiche in un unico DataFrame
    combined_stats = pd.concat([numerical_stats, categorical_stats], axis=1)
    return combined_stats

def save_statistics_to_csv(description: pd.DataFrame, output_csv_path: str):
    description.to_csv(output_csv_path, float_format="%.2f")

def save_statistics_to_excel(description: pd.DataFrame, output_excel_path: str):
    description.to_excel(output_excel_path, engine='openpyxl')
    apply_excel_formatting(output_excel_path)

def apply_excel_formatting(output_excel_path: str):
    wb = load_workbook(output_excel_path)
    ws = wb.active

    # Applica l'allineamento del testo e auto-dimensionamento delle colonne
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Auto-dimensionamento delle colonne
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    wb.save(output_excel_path)

def generate_charts(dataset: pd.DataFrame, output_charts_path: str):
    os.makedirs(output_charts_path, exist_ok=True)

    # Istogramma per release_year
    plt.figure(figsize=(10, 6))
    dataset['release_year'].dropna().astype(int).hist(bins=30, edgecolor='black', color='skyblue')
    plt.title('Distribuzione degli Anni di Rilascio')
    plt.xlabel('Anno di Rilascio')
    plt.ylabel('Numero di Film/Serie')
    plt.grid(axis='y', linestyle='--', linewidth=0.7)
    plt.savefig(os.path.join(output_charts_path, 'release_year_distribution.png'))
    plt.close()

    # Grafico a barre per type
    plt.figure(figsize=(10, 6))
    dataset['type'].value_counts().plot(kind='bar', edgecolor='black', color='skyblue')
    plt.title('Distribuzione per Tipo (Movie/TV Show)')
    plt.xlabel('Tipo')
    plt.ylabel('Conteggio')
    plt.grid(axis='y', linestyle='--', linewidth=0.7)
    plt.savefig(os.path.join(output_charts_path, 'type_distribution.png'))
    plt.close()

    # Grafico a barre per rating
    plt.figure(figsize=(10, 6))
    dataset['rating'].value_counts().plot(kind='bar', edgecolor='black', color='skyblue')
    plt.title('Distribuzione dei Rating')
    plt.xlabel('Rating')
    plt.ylabel('Conteggio')
    plt.grid(axis='y', linestyle='--', linewidth=0.7)
    plt.savefig(os.path.join(output_charts_path, 'rating_distribution.png'))
    plt.close()

# Esegui l'analisi
file_path = os.path.join(os.path.dirname(__file__), '..', 'source', 'amazon_prime_titles.csv')
output_csv_path = os.path.join(os.path.dirname(__file__), '..', 'source', 'statistics.csv')
output_excel_path = os.path.join(os.path.dirname(__file__), '..', 'source', 'statistics_excel.xlsx')
output_charts_path = os.path.join(os.path.dirname(__file__), '..', 'charts', 'analyze_data')

try:
    dataset = load_dataset(file_path)
    dataset = convert_to_categorical(dataset, ['type', 'rating', 'listed_in'])
    
    description = generate_statistics(dataset)
    save_statistics_to_csv(description, output_csv_path)
    save_statistics_to_excel(description, output_excel_path)
    
    generate_charts(dataset, output_charts_path)

    print(f"Analisi completata. Le statistiche sono state salvate in '{output_csv_path}' e '{output_excel_path}'. Le visualizzazioni grafiche sono state salvate nella cartella '{output_charts_path}'.")
    
except Exception as e:
    print(f"Errore durante l'analisi: {e}")
